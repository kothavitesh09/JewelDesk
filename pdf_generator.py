from io import BytesIO
import os
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple
from urllib.parse import urlparse
from urllib.request import urlopen

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

from config import BILL_TEMPLATE_PATH, SHOP_ADDRESS, SHOP_GSTIN, SHOP_NAME, SHOP_PHONE
from utils import format_date_for_pdf, format_invoice_no, rupees_in_words


TEMPLATE_SHEET_MAX_COL = 10
TEMPLATE_SHEET_MAX_ROW = 38
EMU_PER_POINT = 12700.0


def _money(value: Any) -> str:
    try:
        return f"{float(value):,.2f}"
    except Exception:
        return "0.00"


def _qty(value: Any) -> str:
    try:
        num = float(value)
    except Exception:
        return ""
    if num == 0:
        return ""
    return f"{num:,.3f}".rstrip("0").rstrip(".")


def _rate(value: Any) -> str:
    try:
        return f"{float(value):,.2f}"
    except Exception:
        return ""


def _find_template_path() -> Path:
    configured = Path(BILL_TEMPLATE_PATH)
    if not configured.is_absolute():
        configured = Path(__file__).resolve().parent / configured

    template_path_candidates = (
        configured,
        Path(__file__).with_name("Billing Format.xlsx"),
        Path(r"C:\Users\Kotha Vitesh\Downloads\Billing Format.xlsx"),
    )
    for path in template_path_candidates:
        if path.exists():
            return path
    raise FileNotFoundError("Billing Format.xlsx template not found.")


def _load_template_sheet():
    workbook = load_workbook(_find_template_path())
    return workbook.active


def _column_widths(ws) -> List[float]:
    widths: List[float] = []
    for idx in range(1, TEMPLATE_SHEET_MAX_COL + 1):
        letter = chr(64 + idx)
        widths.append(float(ws.column_dimensions[letter].width or 10.0))
    return widths


def _row_heights(ws) -> List[float]:
    heights: List[float] = []
    for idx in range(1, TEMPLATE_SHEET_MAX_ROW + 1):
        heights.append(float(ws.row_dimensions[idx].height or 18.0))
    return heights


def _column_widths_range(ws, start_col: int, end_col: int) -> List[float]:
    widths: List[float] = []
    for idx in range(start_col, end_col + 1):
        letter = chr(64 + idx)
        widths.append(float(ws.column_dimensions[letter].width or 10.0))
    return widths


def _row_heights_range(ws, start_row: int, end_row: int) -> List[float]:
    heights: List[float] = []
    for idx in range(start_row, end_row + 1):
        heights.append(float(ws.row_dimensions[idx].height or 18.0))
    return heights


def _positions(lengths: Iterable[float], start: float, total_span: float) -> List[float]:
    source = list(lengths)
    scale = total_span / sum(source)
    result = [start]
    current = start
    for value in source:
        current += value * scale
        result.append(current)
    return result


def _merge_map(ws) -> Dict[str, Tuple[int, int, int, int]]:
    mapping: Dict[str, Tuple[int, int, int, int]] = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = ws.cell(min_row, min_col).coordinate
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                mapping[ws.cell(row, col).coordinate] = (min_row, min_col, max_row, max_col)
        mapping[top_left] = (min_row, min_col, max_row, max_col)
    return mapping


def _box_for(row: int, col: int, row_edges: List[float], col_edges: List[float], merge: Tuple[int, int, int, int] | None = None):
    min_row, min_col, max_row, max_col = merge or (row, col, row, col)
    left = col_edges[min_col - 1]
    right = col_edges[max_col]
    top = row_edges[min_row - 1]
    bottom = row_edges[max_row]
    return left, bottom, right, top


def _box_for_range(
    row: int,
    col: int,
    row_edges: List[float],
    col_edges: List[float],
    row_start: int,
    col_start: int,
    merge: Tuple[int, int, int, int] | None = None,
):
    min_row, min_col, max_row, max_col = merge or (row, col, row, col)
    left = col_edges[min_col - col_start]
    right = col_edges[(max_col - col_start) + 1]
    top = row_edges[min_row - row_start]
    bottom = row_edges[(max_row - row_start) + 1]
    return left, bottom, right, top


def _font_name(cell) -> str:
    return "Helvetica-Bold" if cell.font and cell.font.bold else "Helvetica"


def _font_size(cell) -> float:
    size = float(cell.font.sz or 10.0) if cell.font else 10.0
    return max(7.0, min(size, 14.0))


def _wrap_lines(text: str, font_name: str, font_size: float, max_width: float, allow_wrap: bool) -> List[str]:
    parts: List[str] = []
    for raw_line in str(text or "").splitlines() or [""]:
        if not allow_wrap:
            parts.append(raw_line)
            continue
        words = raw_line.split()
        if not words:
            parts.append("")
            continue
        line = words[0]
        for word in words[1:]:
            candidate = f"{line} {word}"
            if stringWidth(candidate, font_name, font_size) <= max_width:
                line = candidate
            else:
                parts.append(line)
                line = word
        parts.append(line)
    return parts


def _draw_text_in_box(c: canvas.Canvas, text: str, box, cell, padding: float = 3.5):
    if text in (None, ""):
        return

    left, bottom, right, top = box
    font_name = _font_name(cell)
    font_size = _font_size(cell)
    align = (cell.alignment.horizontal or "left") if cell.alignment else "left"
    vertical = (cell.alignment.vertical or "center") if cell.alignment else "center"
    max_width = max(10, right - left - (padding * 2))
    max_height = max(8, top - bottom - (padding * 2))
    allow_wrap = bool(getattr(cell.alignment, "wrap_text", False)) or ("\n" in str(text))

    lines: List[str] = []
    line_height = 0.0
    total_height = 0.0
    while font_size >= 6.0:
        lines = _wrap_lines(str(text), font_name, font_size, max_width, allow_wrap)
        widest = max((stringWidth(line, font_name, font_size) for line in lines), default=0)
        line_height = font_size * 1.15
        total_height = line_height * len(lines)
        if widest <= max_width and total_height <= max_height:
            break
        font_size -= 0.4

    if vertical == "top":
        y = top - padding - font_size
    elif vertical == "bottom":
        y = bottom + padding + total_height - line_height
    else:
        y = bottom + ((top - bottom) + total_height) / 2 - line_height

    c.setFont(font_name, font_size)
    for line in lines:
        width = stringWidth(line, font_name, font_size)
        if align == "center":
            x = left + (right - left - width) / 2
        elif align == "right":
            x = right - padding - width
        else:
            x = left + padding
        c.drawString(x, y, line)
        y -= line_height


def _draw_border_side(c: canvas.Canvas, side, x1: float, y1: float, x2: float, y2: float):
    if not side or not side.style:
        return

    if side.style == "double":
        line_width = 0.8
        offset = 1.1
        c.setLineWidth(line_width)
        if abs(y1 - y2) < 0.01:
            c.line(x1, y1 - offset, x2, y2 - offset)
            c.line(x1, y1 + offset, x2, y2 + offset)
        else:
            c.line(x1 - offset, y1, x2 - offset, y2)
            c.line(x1 + offset, y1, x2 + offset, y2)
        return

    widths = {
        "hair": 0.3,
        "thin": 0.6,
        "medium": 1.0,
        "thick": 1.4,
    }
    c.setLineWidth(widths.get(side.style, 0.6))
    c.line(x1, y1, x2, y2)


def _draw_borders(c: canvas.Canvas, ws, row_edges: List[float], col_edges: List[float]):
    merges = _merge_map(ws)

    def same_merged_block(row_a: int, col_a: int, row_b: int, col_b: int) -> bool:
        coord_a = ws.cell(row_a, col_a).coordinate
        coord_b = ws.cell(row_b, col_b).coordinate
        merge_a = merges.get(coord_a)
        merge_b = merges.get(coord_b)
        return merge_a is not None and merge_a == merge_b

    for row in range(1, TEMPLATE_SHEET_MAX_ROW + 1):
        for col in range(1, TEMPLATE_SHEET_MAX_COL + 1):
            cell = ws.cell(row, col)
            left, bottom, right, top = _box_for(row, col, row_edges, col_edges)
            if col == 1 or not same_merged_block(row, col, row, col - 1):
                _draw_border_side(c, cell.border.left, left, bottom, left, top)
            if col == TEMPLATE_SHEET_MAX_COL or not same_merged_block(row, col, row, col + 1):
                _draw_border_side(c, cell.border.right, right, bottom, right, top)
            if row == 1 or not same_merged_block(row, col, row - 1, col):
                _draw_border_side(c, cell.border.top, left, top, right, top)
            if row == TEMPLATE_SHEET_MAX_ROW or not same_merged_block(row, col, row + 1, col):
                _draw_border_side(c, cell.border.bottom, left, bottom, right, bottom)


def _amount_to_words_under_1000(value: int) -> str:
    ones = [
        "Zero", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
        "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
        "Seventeen", "Eighteen", "Nineteen",
    ]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    if value < 20:
        return ones[value]
    if value < 100:
        return tens[value // 10] + ("" if value % 10 == 0 else f" {ones[value % 10]}")
    return f"{ones[value // 100]} Hundred" + ("" if value % 100 == 0 else f" {_amount_to_words_under_1000(value % 100)}")


def _amount_to_words(value: float) -> str:
    try:
        return rupees_in_words(float(value))
    except Exception:
        return ""


def _shop_value(shop_overrides: Dict[str, str] | None, invoice: Dict[str, Any], key: str, default: str = "") -> str:
    if shop_overrides and shop_overrides.get(key):
        return str(shop_overrides.get(key) or "").strip()
    return str(invoice.get(key) or default).strip()


def _split_shop_address(address: str) -> List[str]:
    parts = [part.strip() for part in str(address or "").replace("\r", "\n").split("\n") if part.strip()]
    if len(parts) == 1:
        parts = [part.strip() for part in str(address or "").split(",") if part.strip()]
    lines = parts[:3]
    while len(lines) < 3:
        lines.append("")
    return lines


def _dynamic_cell_values(
    invoice: Dict[str, Any],
    items: List[Dict[str, Any]],
    last_page: bool,
    shop_overrides: Dict[str, str] | None = None,
) -> Dict[str, str]:
    invoice_no_text = invoice.get("invoice_no_text")
    if not invoice_no_text and invoice.get("invoice_no") is not None:
        try:
            invoice_no_text = format_invoice_no(int(invoice.get("invoice_no")))
        except Exception:
            invoice_no_text = str(invoice.get("invoice_no"))
    shop_address_lines = _split_shop_address(_shop_value(shop_overrides, invoice, "shop_address", SHOP_ADDRESS))

    dynamic: Dict[str, str] = {
        "I1": f"GSTIN: {_shop_value(shop_overrides, invoice, 'shop_gstin', SHOP_GSTIN)}",
        "I2": shop_address_lines[0],
        "I3": shop_address_lines[1],
        "I4": shop_address_lines[2],
        "I5": f"Call: {_shop_value(shop_overrides, invoice, 'shop_phone', SHOP_PHONE)}",
        "A6": f"Name: {str(invoice.get('customer_name') or '').strip()}",
        "A7": f"Address: {str(invoice.get('customer_address') or '').strip()}",
        "A8": f"Phone: {str(invoice.get('customer_phone') or '').strip()}",
        "A9": f"GSTIN No: {str(invoice.get('party_gst_no') or '').strip()}",
        "I7": "Bill No:",
        "J7": str(invoice_no_text or ""),
        "I8": "Date:",
        "J8": format_date_for_pdf(invoice["created_at"]) if invoice.get("created_at") else "",
    }

    for idx, item in enumerate(items, start=11):
        dynamic[f"A{idx}"] = str(idx - 10)
        dynamic[f"B{idx}"] = str(item.get("particulars") or "")
        dynamic[f"C{idx}"] = _qty(item.get("quantity"))
        dynamic[f"D{idx}"] = _qty(item.get("gross_weight"))
        dynamic[f"E{idx}"] = _qty(item.get("stone_weight"))
        dynamic[f"F{idx}"] = _qty(item.get("qty_gms"))
        dynamic[f"G{idx}"] = _qty(item.get("value_addition"))
        dynamic[f"H{idx}"] = _rate(item.get("rate_per_g"))
        dynamic[f"I{idx}"] = _money(item.get("stone_amount", 0))
        dynamic[f"J{idx}"] = _money(item.get("invoice_amount", item.get("amount", 0)))

    if last_page:
        dynamic["J26"] = _money(invoice.get("total", 0))
        dynamic["J27"] = _money(invoice.get("discount", 0))
        dynamic["J28"] = _money(invoice.get("sgst", 0))
        dynamic["J29"] = _money(invoice.get("cgst", 0))
        dynamic["J30"] = _money(invoice.get("igst", 0))
        dynamic["J31"] = _money(invoice.get("final_amount", 0))
        dynamic["C32"] = _amount_to_words(invoice.get("final_amount", 0))
        dynamic["F34"] = f"For {_shop_value(shop_overrides, invoice, 'shop_name', SHOP_NAME) or SHOP_NAME}"
    else:
        dynamic["C32"] = "Continued on next page..."

    return dynamic


def _resolve_image_source(path_value: str | None, fallback: str | None = None):
    for candidate in (path_value, fallback):
        if not candidate:
            continue
        candidate_str = str(candidate).strip()
        parsed = urlparse(candidate_str)
        if parsed.scheme in {"http", "https"} and parsed.netloc:
            return candidate_str
        path = Path(candidate_str)
        if not path.is_absolute():
            path = Path(__file__).resolve().parent / path
        if path.exists() and path.is_file():
            return path
    return None


def _anchor_box(anchor, row_edges: List[float], col_edges: List[float]):
    from_marker = anchor._from
    to_marker = anchor.to

    left = col_edges[from_marker.col] + (from_marker.colOff / EMU_PER_POINT)
    right = col_edges[to_marker.col] + (to_marker.colOff / EMU_PER_POINT)
    top = row_edges[from_marker.row] - (from_marker.rowOff / EMU_PER_POINT)
    bottom = row_edges[to_marker.row] - (to_marker.rowOff / EMU_PER_POINT)
    return left, bottom, right, top


def _anchor_box_in_range(anchor, row_edges: List[float], col_edges: List[float], row_start: int, col_start: int):
    from_marker = anchor._from
    to_marker = anchor.to

    left = col_edges[from_marker.col - (col_start - 1)] + (from_marker.colOff / EMU_PER_POINT)
    right = col_edges[to_marker.col - (col_start - 1)] + (to_marker.colOff / EMU_PER_POINT)
    top = row_edges[from_marker.row - (row_start - 1)] - (from_marker.rowOff / EMU_PER_POINT)
    bottom = row_edges[to_marker.row - (row_start - 1)] - (to_marker.rowOff / EMU_PER_POINT)
    return left, bottom, right, top


def _draw_image(c: canvas.Canvas, image_source, box):
    if image_source is None:
        return
    try:
        left, bottom, right, top = box
        if right <= left or top <= bottom:
            return
        if isinstance(image_source, Path):
            img = ImageReader(str(image_source))
        else:
            with urlopen(str(image_source)) as response:
                img = ImageReader(BytesIO(response.read()))
        c.drawImage(img, left, bottom, width=right - left, height=top - bottom, preserveAspectRatio=True, anchor="c", mask="auto")
    except Exception:
        return


def _draw_shop_name_text(c: canvas.Canvas, ws, box, shop_name: str):
    if not shop_name:
        return
    try:
        cell = ws["B1"]
        _draw_text_in_box(c, shop_name, box, cell)
    except Exception:
        return


def _draw_header_images(c: canvas.Canvas, ws, row_edges: List[float], col_edges: List[float], invoice: Dict[str, Any], shop_overrides: Dict[str, str] | None = None):
    images = sorted(
        [img for img in getattr(ws, "_images", []) if getattr(img.anchor, "_from", None) and img.anchor._from.row < 5],
        key=lambda img: (img.anchor._from.row, img.anchor._from.col),
    )

    logo_box = _anchor_box(images[0].anchor, row_edges, col_edges) if len(images) >= 1 else _box_for(1, 1, row_edges, col_edges, (1, 1, 5, 2))
    shop_name_box = _anchor_box(images[1].anchor, row_edges, col_edges) if len(images) >= 2 else _box_for(1, 2, row_edges, col_edges, (1, 2, 5, 8))

    logo_path = _resolve_image_source(
        _shop_value(shop_overrides, invoice, "logo_path"),
        os.path.join("static", "images", "shop_logo.png"),
    )
    shop_name_image_path = _resolve_image_source(
        _shop_value(shop_overrides, invoice, "shop_name_image_path"),
        os.path.join("static", "images", "shop_name.png"),
    )
    shop_name = _shop_value(shop_overrides, invoice, "shop_name", SHOP_NAME) or SHOP_NAME

    _draw_image(c, logo_path, logo_box)
    if shop_name_image_path is not None:
        _draw_image(c, shop_name_image_path, shop_name_box)
    else:
        _draw_shop_name_text(c, ws, shop_name_box, shop_name)


def _draw_terms_page_image(c: canvas.Canvas, ws, row_edges: List[float], col_edges: List[float], invoice: Dict[str, Any], shop_overrides: Dict[str, str] | None = None):
    image = next(
        (
            img
            for img in getattr(ws, "_images", [])
            if getattr(img.anchor, "_from", None)
            and img.anchor._from.row == 40
            and img.anchor._from.col == 1
        ),
        None,
    )
    if image is None:
        return

    logo_path = _resolve_image_source(
        _shop_value(shop_overrides, invoice, "logo_path"),
        os.path.join("static", "images", "shop_logo.png"),
    )
    _draw_image(c, logo_path, _anchor_box_in_range(image.anchor, row_edges, col_edges, 41, 2))


def _render_page(
    c: canvas.Canvas,
    ws,
    invoice: Dict[str, Any],
    items: List[Dict[str, Any]],
    last_page: bool,
    shop_overrides: Dict[str, str] | None = None,
):
    page_w, page_h = A4
    left_margin = 8 * mm
    right_margin = 8 * mm
    top_margin = 12 * mm
    bottom_margin = 10 * mm
    usable_w = page_w - left_margin - right_margin
    usable_h = page_h - top_margin - bottom_margin

    col_edges = _positions(_column_widths(ws), left_margin, usable_w)
    y_positions = _positions(_row_heights(ws), bottom_margin, usable_h)
    row_edges = [page_h - y for y in y_positions]

    c.setLineJoin(1)
    c.setLineCap(1)
    c.setStrokeColorRGB(0, 0, 0)
    c.setFillColorRGB(0, 0, 0)

    _draw_borders(c, ws, row_edges, col_edges)

    merges = _merge_map(ws)
    skip_rows = set() if last_page else set(range(26, 39))
    dynamic = _dynamic_cell_values(invoice, items, last_page, shop_overrides=shop_overrides)

    for row in range(1, TEMPLATE_SHEET_MAX_ROW + 1):
        for col in range(1, TEMPLATE_SHEET_MAX_COL + 1):
            cell = ws.cell(row, col)
            coord = cell.coordinate
            merge = merges.get(coord)
            if merge:
                min_row, min_col, _, _ = merge
                if row != min_row or col != min_col:
                    continue

            if row in skip_rows and coord not in dynamic:
                continue

            box = _box_for(row, col, row_edges, col_edges, merge)
            text = dynamic.get(coord, cell.value)
            _draw_text_in_box(c, text, box, cell)

    _draw_header_images(c, ws, row_edges, col_edges, invoice, shop_overrides=shop_overrides)

    if not last_page:
        notice_box = _box_for(32, 3, row_edges, col_edges, (32, 3, 32, 10))
        temp_cell = ws["C32"]
        _draw_text_in_box(c, "Continued on next page", notice_box, temp_cell)


def _render_terms_page(c: canvas.Canvas, ws, invoice: Dict[str, Any], shop_overrides: Dict[str, str] | None = None):
    page_w, page_h = A4
    left_margin = 8 * mm
    right_margin = 8 * mm
    top_margin = 12 * mm
    bottom_margin = 10 * mm
    usable_w = page_w - left_margin - right_margin
    usable_h = page_h - top_margin - bottom_margin

    col_start = 2
    col_end = 10
    row_start = 41
    row_end = 67

    col_edges = _positions(_column_widths_range(ws, col_start, col_end), left_margin, usable_w)
    y_positions = _positions(_row_heights_range(ws, row_start, row_end), bottom_margin, usable_h)
    row_edges = [page_h - y for y in y_positions]

    c.setLineJoin(1)
    c.setLineCap(1)
    c.setStrokeColorRGB(0, 0, 0)
    c.setFillColorRGB(0, 0, 0)

    merges = _merge_map(ws)
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row, col)
            coord = cell.coordinate
            merge = merges.get(coord)
            if merge:
                min_row, min_col, _, _ = merge
                if row != min_row or col != min_col:
                    continue
            box = _box_for_range(row, col, row_edges, col_edges, row_start, col_start, merge)
            _draw_text_in_box(c, cell.value, box, cell)

    _draw_terms_page_image(c, ws, row_edges, col_edges, invoice, shop_overrides=shop_overrides)


def generate_invoice_pdf_bytes(invoice: Dict[str, Any], shop_overrides: Dict[str, str] | None = None) -> bytes:
    ws = _load_template_sheet()
    items = list(invoice.get("items") or [])
    pages = [items[i:i + 15] for i in range(0, len(items), 15)] or [[]]

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)

    for idx, page_items in enumerate(pages):
        _render_page(c, ws, invoice, page_items, last_page=(idx == len(pages) - 1), shop_overrides=shop_overrides)
        c.showPage()

    if ws.max_row >= 67:
        _render_terms_page(c, ws, invoice, shop_overrides=shop_overrides)
        c.showPage()

    c.save()
    buffer.seek(0)
    return buffer.read()
